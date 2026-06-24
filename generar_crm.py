#!/usr/bin/env python3
"""
Generador de CRM — JA Abogados
Produce crm_jaabogados.xlsx listo para subir a Google Sheets
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from datetime import datetime

# ─── PALETA DE COLORES ────────────────────────────────────────────────────────
NAVY       = "1A3A5C"
GOLD       = "C9A84C"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MED_GRAY   = "D0D0D0"
DARK_GRAY  = "555555"

C_VERDE    = "C6EFCE"  # fondo verde suave
C_VERDE_F  = "276221"  # texto verde oscuro
C_AMARILLO = "FFEB9C"
C_AMARILLO_F = "9C6500"
C_NARANJA  = "FFCCAA"
C_NARANJA_F  = "974706"
C_ROJO     = "FFC7CE"
C_ROJO_F   = "9C0006"
C_HEADER   = NAVY
C_HEADER_F = WHITE

# ─── DATOS: 17 PROSPECTOS ─────────────────────────────────────────────────────
PROSPECTOS = [
    # (empresa, sector, señal, decisor, cargo, email, telefono, linkedin,
    #  servicio, c1,c2,c3,c4,c5,c6,c7,c8, estado, proxima_accion, notas)
    (
        "Telepatía", "IA / Tecnología",
        "Ronda US$42M de a16z (2025)",
        "Tomás Giraldo", "CEO / Co-fundador",
        "info@telepatia.ai", "No disponible",
        "linkedin.com/search/?keywords=Tomás+Giraldo+Telepatía",
        "Contratos · Compliance · Laboral",
        5,5,5,5,4,5,4,3,
        "Prospecto Nuevo", "Enviar conexión LinkedIn esta semana", ""
    ),
    (
        "Recurly Colombia", "SaaS / Facturación",
        "Nueva entidad legal Colombia jun-2024",
        "CFO / Director Jurídico", "Por confirmar",
        "info@recurly.com", "304 588 1010",
        "linkedin.com/company/recurly",
        "Constitución · Contratos · Tributario",
        5,5,5,3,5,5,4,3,
        "Prospecto Nuevo", "Enviar email + LinkedIn esta semana", ""
    ),
    (
        "Health Prime Colombia", "Salud / BPO",
        "Llegada nueva al mercado colombiano 2025",
        "Gerente General", "Por confirmar",
        "No encontrado", "604 444 8883",
        "linkedin.com/company/health-prime-international",
        "Laboral · Contratos · Asesoría empresa extranjera",
        5,5,4,3,5,5,4,3,
        "Prospecto Nuevo", "Llamar + buscar decisor en LinkedIn", ""
    ),
    (
        "CapitalPocket / Pocki", "Fintech",
        "Operación regulada ante Superfinanciera",
        "Alex Aronategui Herrera", "CEO",
        "No encontrado", "No disponible",
        "linkedin.com/in/alex-aronategui-herrera",
        "Compliance financiero · Contratos",
        5,5,4,5,4,5,4,2,
        "Prospecto Nuevo", "Conectar con Alex en LinkedIn", ""
    ),
    (
        "Drones Sky Solutions", "Tecnología Aeronáutica",
        "Regulación Aerocivil activa para drones comerciales",
        "Director General", "Por confirmar",
        "info@dronesskysolutions.com", "+57 319 731 6797",
        "linkedin.com/company/drones-sky-solutions",
        "Compliance · Contratos · Redacción jurídica",
        5,5,4,3,4,5,5,2,
        "Prospecto Nuevo", "Email + WhatsApp esta semana", ""
    ),
    (
        "Netux", "Tecnología / Salud Digital",
        "Seleccionada Medellín Next 2026 — top 5",
        "Gerente General", "Por confirmar",
        "info@netux.com", "No disponible",
        "linkedin.com/company/netux",
        "Laboral · Contratos TI · Corporativo",
        4,4,4,3,4,5,3,3,
        "Prospecto Nuevo", "Mensaje LinkedIn nurturing 30 días", ""
    ),
    (
        "MovilBox", "Desarrollo Móvil / Apps",
        "Seleccionada Medellín Next 2026",
        "Gerente / Fundador", "Por confirmar",
        "sales@movilbox.net", "604 604 0563",
        "linkedin.com/company/movilbox",
        "Propiedad intelectual · Contratos · Laboral",
        4,4,4,3,4,5,3,3,
        "Prospecto Nuevo", "Email sales@movilbox.net esta semana", ""
    ),
    (
        "Medvision", "Medtech / Dispositivos Médicos",
        "Regulación INVIMA activa — dispositivos clase II/III",
        "Director Comercial", "Por confirmar",
        "soporte@medvision.com.co", "+57 300 412 8551",
        "linkedin.com/company/medvision",
        "Compliance INVIMA · Contratos · Redacción",
        4,4,4,3,4,5,4,2,
        "Prospecto Nuevo", "Email + LinkedIn nurturing", ""
    ),
    (
        "Erco Energía", "Energías Renovables",
        "Seleccionada Medellín Next 2026 — top 5",
        "Gerente General", "Por confirmar",
        "No encontrado", "604 604 6391",
        "linkedin.com/company/erco-energia",
        "Contratos B2B · Laboral · Corporativo",
        4,4,4,3,4,5,3,2,
        "Prospecto Nuevo", "Llamar 604 604 6391 — pedir gerente", ""
    ),
    (
        "Golden Energy", "Energía Solar",
        "Expansión proyectos solares en Antioquia",
        "Director de Proyectos", "Por confirmar",
        "proyectos@goldenenergysas.com", "+57 313 566 3288",
        "linkedin.com/search/?keywords=Golden+Energy+Colombia",
        "Contratos · Constitución SAS · Laboral",
        4,4,3,3,4,5,3,2,
        "Prospecto Nuevo", "Email proyectos@goldenenergysas.com", ""
    ),
    (
        "Somos Internet", "Telecomunicaciones",
        "Expansión México con inversión US$40M",
        "CEO / Gerente General", "Por confirmar",
        "No encontrado", "No disponible",
        "linkedin.com/company/somos-internet",
        "Contratos internacionales · Laboral · Expansión",
        4,4,4,2,4,5,3,3,
        "Prospecto Nuevo", "Buscar decisor LinkedIn — seguimiento pasivo", ""
    ),
    (
        "SiembraViva", "Agritech / B-Corp",
        "Crecimiento mercado orgánico — certificación B-Corp",
        "Gerente General", "Por confirmar",
        "No encontrado", "302 213 1321",
        "linkedin.com/company/siembraviva",
        "Contratos · Compliance B-Corp · Laboral",
        3,3,3,2,4,5,3,2,
        "Seguimiento Pasivo", "Conectar LinkedIn — sin prioridad inmediata", ""
    ),
    (
        "Azimut Energía", "Energías Renovables",
        "Seleccionada Medellín Next 2026",
        "Gerente / Fundador", "Por confirmar",
        "No encontrado", "No disponible",
        "linkedin.com/search/?keywords=Azimut+Energía+Medellín",
        "Contratos · Laboral · Corporativo",
        3,3,3,2,4,5,3,1,
        "Seguimiento Pasivo", "Monitorear señales — contactar en 45 días", ""
    ),
    (
        "B2Chat", "SaaS / Comunicaciones",
        "Seleccionada Medellín Next 2026",
        "CEO", "Por confirmar",
        "No encontrado", "314 656 7290",
        "linkedin.com/company/b2chat",
        "Contratos SaaS · Laboral",
        3,3,3,2,3,5,2,2,
        "Seguimiento Pasivo", "Conectar LinkedIn — largo plazo", ""
    ),
    (
        "Aba Tech", "Tecnología",
        "Crecimiento sostenido desde 2022",
        "Gerente General", "Por confirmar",
        "No encontrado", "301 294 7758",
        "linkedin.com/search/?keywords=Aba+Tech+Medellín",
        "Corporativo · Contratos · Laboral",
        3,3,3,2,3,5,2,1,
        "Seguimiento Pasivo", "Monitorear — revisar en 60 días", ""
    ),
    (
        "iQor Colombia", "BPO / Contact Center",
        "Nueva oficina en Medellín 2024",
        "Country Manager Colombia", "Por confirmar",
        "No encontrado", "300 690 1100",
        "linkedin.com/company/iqor",
        "Laboral · Compliance multinacional",
        3,3,4,2,3,5,2,1,
        "Seguimiento Pasivo", "Bajo potencial — posible abogado interno", ""
    ),
    (
        "TaskUs Colombia", "BPO / Outsourcing",
        "4,000+ empleos — expansión Latte Valley",
        "Country Manager Colombia", "Por confirmar",
        "No encontrado", "No disponible",
        "linkedin.com/company/taskus",
        "Laboral",
        3,2,4,2,2,5,2,0,
        "Seguimiento Pasivo", "Muy bajo potencial — probable abogado interno", ""
    ),
]

CRITERIOS = [
    "C1: Señal reciente",
    "C2: Alineación servicios",
    "C3: Tamaño / capacidad pago",
    "C4: Acceso al decisor",
    "C5: Madurez legal empresa",
    "C6: Presencia Medellín",
    "C7: Urgencia legal",
    "C8: Sin riesgo BigLaw",
]

ESTADOS = [
    "Prospecto Nuevo",
    "Contactado",
    "En Conversación",
    "Propuesta Enviada",
    "Cliente Activo",
    "Descartado",
    "Seguimiento Pasivo",
]

MENSAJES = {
    "Telepatía": {
        "conexion": "Hola Tomás, sigo con interés el crecimiento de Telepatía — la ronda de a16z es una señal poderosa. Me dedico al derecho corporativo en Medellín y me encantaría conectar.",
        "mensaje1": "Hola Tomás, gracias por conectar. Felicitaciones por la ronda de inversión — estructurar contratos con inversores internacionales como a16z implica varios puntos críticos de compliance colombiano que muchas startups pasan por alto en esa etapa. Si en algún momento necesitan revisar cláusulas de gobierno corporativo o acuerdos de empleados clave, con gusto les doy una primera revisión sin costo. ¿Están gestionando esto internamente o con apoyo externo?",
        "seguimiento": "Tomás, ¿cómo avanza la estructuración post-ronda? Acabo de publicar una guía sobre las 5 cláusulas que toda startup colombiana debe revisar antes de cerrar una ronda Serie A — se la comparto si le es útil.",
    },
    "Recurly Colombia": {
        "conexion": "Hola, vi que Recurly abrió operaciones en Colombia en 2024. Me especializo en asesoría jurídica para empresas extranjeras que ingresan al mercado colombiano. Me encantaría conectar.",
        "mensaje1": "Buen día. Noté que Recurly constituyó su entidad legal en Colombia el año pasado — un paso importante que trae consigo obligaciones tributarias, laborales y contractuales específicas que difieren del marco estadounidense. Desde mi práctica en derecho corporativo en Medellín, he acompañado a varias empresas SaaS en esa estructuración inicial. ¿Estarían abiertos a una revisión gratuita de su situación de compliance local?",
        "seguimiento": "¿Cómo van adaptando sus contratos de suscripción al marco colombiano? La Ley 1480 (Estatuto del Consumidor) tiene implicaciones específicas para SaaS que vale la pena revisar. Le comparto un resumen si le interesa.",
    },
    "Health Prime Colombia": {
        "conexion": "Hola, sé que Health Prime está iniciando operaciones en Colombia. Me especializo en asesoría a empresas extranjeras del sector salud en Medellín — me encantaría conectar.",
        "mensaje1": "Buen día. Entiendo que Health Prime está estableciendo su operación en Colombia — un mercado con un marco normativo de salud bastante específico (MIPRES, habilitación, protección de datos de pacientes). Desde mi práctica en derecho corporativo y laboral en Medellín, acompaño a empresas del sector en su estructuración local. ¿Tienen cubierta la parte jurídica de la operación colombiana o están buscando apoyo externo?",
        "seguimiento": "Una consulta: ¿ya adelantaron la revisión de sus contratos de prestación de servicios con las IPS colombianas? Hay cláusulas específicas que exige la normativa local que suelen sorprender a empresas con sede en EE.UU. Con gusto les comparto un checklist.",
    },
    "CapitalPocket / Pocki": {
        "conexion": "Hola Alex, sigo el desarrollo de CapitalPocket con mucho interés — el espacio de fintech regulado en Colombia es fascinante. Me dedico al derecho financiero y compliance. Me encantaría conectar.",
        "mensaje1": "Hola Alex, gracias por conectar. Operar en el ecosistema regulado por la Superfinanciera implica un nivel de compliance documental que evoluciona rápido — especialmente con las nuevas circulares de 2025 sobre actividades de crédito digital. Desde mi práctica en Medellín acompaño a fintechs en la estructuración de sus contratos con usuarios y en la revisión de sus manuales de riesgo. ¿Estarían abiertos a conversar sobre cómo tienen estructurado ese frente?",
        "seguimiento": "Alex, ¿cómo les ha ido con la Circular 026 de la SFC sobre protección al consumidor financiero? Hay un par de obligaciones contractuales nuevas que muchas fintechs aún no han incorporado. Le comparto una nota si le resulta útil.",
    },
    "Drones Sky Solutions": {
        "conexion": "Hola, sigo de cerca la regulación de drones comerciales en Colombia — un tema que está evolucionando rápido con Aerocivil. Me dedico al derecho corporativo en Medellín y me gustaría conectar.",
        "mensaje1": "Buen día. La regulación de Aerocivil para operadores de drones comerciales (RAC 91 y resoluciones recientes) implica contratos de operación, seguros de responsabilidad y cláusulas específicas con clientes corporativos que muchas empresas del sector aún no tienen formalizados. Desde mi práctica en Medellín apoyo a empresas de tecnología en esa estructuración. ¿Tienen cubierto ese frente jurídico o les sería útil una revisión inicial?",
        "seguimiento": "¿Cómo estructuran sus contratos con clientes corporativos que contratan levantamientos fotogramétricos o inspecciones? Hay cláusulas de limitación de responsabilidad que son críticas en este sector. Le comparto un modelo si le interesa.",
    },
}

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def header_fill(color=C_HEADER):
    return PatternFill("solid", fgColor=color)

def header_font(color=C_HEADER_F, bold=True, size=11):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def cell_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color=MED_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def semaforo(total):
    if total >= 33: return "🟢 VERDE"
    if total >= 24: return "🟡 AMARILLO"
    if total >= 15: return "🟠 NARANJA"
    return "🔴 ROJO"

def semaforo_fill(total):
    if total >= 33: return PatternFill("solid", fgColor=C_VERDE)
    if total >= 24: return PatternFill("solid", fgColor=C_AMARILLO)
    if total >= 15: return PatternFill("solid", fgColor=C_NARANJA)
    return PatternFill("solid", fgColor=C_ROJO)

def semaforo_font(total):
    if total >= 33: return Font(name="Calibri", bold=True, color=C_VERDE_F, size=10)
    if total >= 24: return Font(name="Calibri", bold=True, color=C_AMARILLO_F, size=10)
    if total >= 15: return Font(name="Calibri", bold=True, color=C_NARANJA_F, size=10)
    return Font(name="Calibri", bold=True, color=C_ROJO_F, size=10)

# ─── HOJA 1: DASHBOARD ───────────────────────────────────────────────────────

def crear_dashboard(wb, prospectos):
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False

    # Título
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "JA ABOGADOS — CRM DE PROSPECTOS JURÍDICOS"
    c.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    c.fill = header_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Medellín, Colombia  |  Actualizado: {datetime.now().strftime('%d/%m/%Y')}"
    c.font = Font(name="Calibri", size=10, color=GOLD)
    c.fill = header_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[2].height = 20

    # Separador
    ws.row_dimensions[3].height = 8

    # Contadores por semáforo
    totals = [sum(p[9:17]) for p in prospectos]
    verde    = sum(1 for t in totals if t >= 33)
    amarillo = sum(1 for t in totals if 24 <= t < 33)
    naranja  = sum(1 for t in totals if 15 <= t < 24)
    rojo     = sum(1 for t in totals if t < 15)
    total    = len(prospectos)

    stats = [
        ("🟢 VERDE\nPrioridad Alta", verde,   C_VERDE,    C_VERDE_F),
        ("🟡 AMARILLO\nNurturing",   amarillo, C_AMARILLO, C_AMARILLO_F),
        ("🟠 NARANJA\nSeg. Pasivo",  naranja,  C_NARANJA,  C_NARANJA_F),
        ("🔴 ROJO\nDescartar",       rojo,     C_ROJO,     C_ROJO_F),
        ("📋 TOTAL\nProspectos",     total,    "E8E8E8",   DARK_GRAY),
    ]

    cols = [1, 2, 3, 4, 5]  # columnas A-E
    ws.row_dimensions[4].height = 14
    ws.row_dimensions[5].height = 40
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 14

    for i, (label, count, bg, fg) in enumerate(stats):
        col = cols[i]
        # Etiqueta
        cl = ws.cell(row=5, column=col)
        cl.value = label
        cl.fill = PatternFill("solid", fgColor=bg)
        cl.font = Font(name="Calibri", bold=True, size=9, color=fg)
        cl.alignment = center()
        cl.border = thin_border()
        # Número
        cn = ws.cell(row=6, column=col)
        cn.value = count
        cn.fill = PatternFill("solid", fgColor=bg)
        cn.font = Font(name="Calibri", bold=True, size=22, color=fg)
        cn.alignment = center()
        cn.border = thin_border()

    # Separador
    ws.row_dimensions[8].height = 14

    # Top 5 prospectos prioritarios
    ws.merge_cells("A9:H9")
    ch = ws["A9"]
    ch.value = "TOP 5 — PROSPECTOS PRIORITARIOS (Semáforo Verde)"
    ch.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    ch.fill = header_fill(NAVY)
    ch.alignment = center()
    ws.row_dimensions[9].height = 22

    hdrs = ["#", "Empresa", "Sector", "Decisor / Cargo", "Servicio Principal", "Puntaje", "Semáforo", "Próxima Acción"]
    for j, h in enumerate(hdrs):
        cc = ws.cell(row=10, column=j+1)
        cc.value = h
        cc.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cc.fill = header_fill(GOLD)
        cc.alignment = center()
        cc.border = thin_border()
    ws.row_dimensions[10].height = 18

    sorted_p = sorted(prospectos, key=lambda x: sum(x[9:17]), reverse=True)
    for idx, p in enumerate(sorted_p[:5]):
        row = 11 + idx
        total_p = sum(p[9:17])
        values = [
            idx+1, p[0], p[1],
            f"{p[3]} — {p[4]}",
            p[8], total_p,
            semaforo(total_p),
            p[19]
        ]
        for j, v in enumerate(values):
            cc = ws.cell(row=row, column=j+1)
            cc.value = v
            cc.font = cell_font(size=9)
            cc.alignment = left() if j in [1,2,3,4,7] else center()
            cc.border = thin_border()
            if j == 6:  # semáforo
                cc.fill = semaforo_fill(total_p)
                cc.font = semaforo_font(total_p)
            elif idx % 2 == 0:
                cc.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        ws.row_dimensions[row].height = 18

    # Anchos de columna
    for col, w in zip("ABCDEFGH", [6, 20, 18, 24, 26, 9, 16, 36]):
        ws.column_dimensions[col].width = w

    # Columnas F-H vacías del bloque stats
    for col in "FGHI":
        ws.column_dimensions[col].width = 5

# ─── HOJA 2: CRM PROSPECTOS ──────────────────────────────────────────────────

def crear_crm(wb, prospectos):
    ws = wb.create_sheet("📋 CRM Prospectos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Título
    ws.merge_cells("A1:Z1")
    c = ws["A1"]
    c.value = "CRM DE PROSPECTOS — JA ABOGADOS  |  Medellín, Colombia"
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = header_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    # Cabeceras
    COLS = [
        ("#", 4),
        ("Empresa", 22),
        ("Sector", 18),
        ("Señal de Oportunidad", 30),
        ("Decisor", 20),
        ("Cargo", 16),
        ("Email", 24),
        ("Teléfono", 14),
        ("LinkedIn", 14),
        ("Canal Contacto", 14),
        ("Fecha 1er Contacto", 14),
        ("ESTADO", 18),
        ("Servicio Principal", 28),
        ("C1\nSeñal", 7),
        ("C2\nServicios", 7),
        ("C3\nTamaño", 7),
        ("C4\nDecisor", 7),
        ("C5\nMadurez", 7),
        ("C6\nMedellín", 7),
        ("C7\nUrgencia", 7),
        ("C8\nBigLaw", 7),
        ("TOTAL\n/40", 8),
        ("SEMÁFORO", 14),
        ("Próxima Acción", 30),
        ("Fecha Próx. Acción", 14),
        ("Notas", 30),
    ]

    for j, (name, width) in enumerate(COLS):
        col_letter = get_column_letter(j+1)
        ws.column_dimensions[col_letter].width = width
        c = ws.cell(row=2, column=j+1)
        c.value = name
        c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        c.fill = header_fill(GOLD) if j in range(13, 22) else header_fill(NAVY)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[2].height = 32

    # Validación de Estado (dropdown)
    estado_formula = '"' + ",".join(ESTADOS) + '"'
    dv = DataValidation(type="list", formula1=estado_formula, allow_blank=True)
    dv.prompt = "Seleccione el estado del prospecto"
    ws.add_data_validation(dv)

    # Validación de scores 0-5
    score_dv = DataValidation(type="whole", operator="between",
                              formula1="0", formula2="5", allow_blank=True)
    score_dv.prompt = "Ingrese un valor entre 0 y 5"
    ws.add_data_validation(score_dv)

    # Canal de contacto (dropdown)
    canal_formula = '"LinkedIn,Email,WhatsApp,Llamada,Referido,Web"'
    canal_dv = DataValidation(type="list", formula1=canal_formula, allow_blank=True)
    ws.add_data_validation(canal_dv)

    # Filas de datos
    for i, p in enumerate(prospectos):
        row = i + 3
        total_p = sum(p[9:17])
        sem = semaforo(total_p)

        values = [
            i+1,          # A: #
            p[0],         # B: Empresa
            p[1],         # C: Sector
            p[2],         # D: Señal
            p[3],         # E: Decisor
            p[4],         # F: Cargo
            p[5],         # G: Email
            p[6],         # H: Teléfono
            p[7],         # I: LinkedIn (texto, no hiperlink para compatibilidad)
            "LinkedIn",   # J: Canal
            "",           # K: Fecha primer contacto
            p[17],        # L: Estado
            p[8],         # M: Servicio
            p[9],         # N: C1
            p[10],        # O: C2
            p[11],        # P: C3
            p[12],        # Q: C4
            p[13],        # R: C5
            p[14],        # S: C6
            p[15],        # T: C7
            p[16],        # U: C8
            total_p,      # V: Total
            sem,          # W: Semáforo
            p[18],        # X: Próxima acción
            "",           # Y: Fecha próxima acción
            p[19],        # Z: Notas
        ]

        bg = LIGHT_GRAY if i % 2 == 0 else WHITE

        for j, v in enumerate(values):
            c = ws.cell(row=row, column=j+1)
            c.value = v
            c.border = thin_border()

            # Formato por columna
            if j == 0:   # #
                c.font = cell_font(bold=True, size=9)
                c.alignment = center()
                c.fill = PatternFill("solid", fgColor=bg)
            elif j in [1, 3, 8, 12, 23, 25]:  # texto largo
                c.font = cell_font(size=9)
                c.alignment = left()
                c.fill = PatternFill("solid", fgColor=bg)
            elif j == 11:  # Estado
                c.font = cell_font(bold=True, size=9)
                c.alignment = center()
                c.fill = PatternFill("solid", fgColor=bg)
                dv.add(c)
            elif j in range(13, 21):  # C1-C8
                c.font = cell_font(bold=True, size=9)
                c.alignment = center()
                score_dv.add(c)
                # Colorear según valor
                if isinstance(v, int):
                    if v >= 4:
                        c.fill = PatternFill("solid", fgColor=C_VERDE)
                        c.font = Font(name="Calibri", bold=True, size=9, color=C_VERDE_F)
                    elif v == 3:
                        c.fill = PatternFill("solid", fgColor=C_AMARILLO)
                        c.font = Font(name="Calibri", bold=True, size=9, color=C_AMARILLO_F)
                    elif v <= 2:
                        c.fill = PatternFill("solid", fgColor=C_NARANJA)
                        c.font = Font(name="Calibri", bold=True, size=9, color=C_NARANJA_F)
            elif j == 21:  # Total
                c.font = Font(name="Calibri", bold=True, size=11, color=NAVY)
                c.alignment = center()
                c.fill = PatternFill("solid", fgColor=bg)
            elif j == 22:  # Semáforo
                c.fill = semaforo_fill(total_p)
                c.font = semaforo_font(total_p)
                c.alignment = center()
            elif j == 9:  # Canal
                canal_dv.add(c)
                c.font = cell_font(size=9)
                c.alignment = center()
                c.fill = PatternFill("solid", fgColor=bg)
            elif j in [10, 24]:  # fechas
                c.font = cell_font(size=9)
                c.alignment = center()
                c.number_format = "DD/MM/YYYY"
                c.fill = PatternFill("solid", fgColor=bg)
            else:
                c.font = cell_font(size=9)
                c.alignment = center()
                c.fill = PatternFill("solid", fgColor=bg)

        ws.row_dimensions[row].height = 20

# ─── HOJA 3: MATRIZ DE EVALUACIÓN ────────────────────────────────────────────

def crear_matriz(wb):
    ws = wb.create_sheet("🎯 Matriz de Evaluación")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "MATRIZ DE VIABILIDAD DE PROSPECTOS — JA ABOGADOS"
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = header_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = "Califica cada criterio de 0 a 5. Puntaje máximo: 40 puntos."
    c.font = Font(name="Calibri", size=10, color=GOLD, italic=True)
    c.fill = header_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 10

    # Tabla de criterios
    hdrs = ["Criterio", "Descripción", "0 puntos", "3 puntos", "5 puntos", "Peso", "Nota"]
    widths = [22, 35, 28, 28, 28, 8, 20]
    for j, (h, w) in enumerate(zip(hdrs, widths)):
        col = get_column_letter(j+1)
        ws.column_dimensions[col].width = w
        c = ws.cell(row=4, column=j+1)
        c.value = h
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = header_fill(GOLD)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[4].height = 20

    criterios_data = [
        (
            "C1: Señal Reciente",
            "¿Hubo un evento reciente que crea necesidad jurídica?",
            "Sin señal visible",
            "Señal indirecta (crecimiento general)",
            "Señal directa: apertura, sanción, inversión, expansión",
            "Alto",
            "Verificar < 12 meses"
        ),
        (
            "C2: Alineación Servicios",
            "¿Cuántos de los 8 servicios de la firma encajan?",
            "Ningún servicio encaja",
            "1–2 servicios encajan",
            "3 o más servicios encajan",
            "Alto",
            "Revisar portafolio"
        ),
        (
            "C3: Tamaño / Capacidad Pago",
            "¿Puede la empresa pagar honorarios de firma externa?",
            "Micro < $50M/año",
            "Pyme $50M–$800M/año",
            "Empresa > $800M/año o financiada",
            "Alto",
            "Estimar por sector"
        ),
        (
            "C4: Acceso al Decisor",
            "¿Se puede contactar al decisor directamente?",
            "Decisor no identificado",
            "Identificado sin canal directo",
            "LinkedIn/email del decisor confirmado",
            "Medio",
            "Buscar en LinkedIn"
        ),
        (
            "C5: Madurez Legal",
            "¿La empresa tiene necesidades jurídicas evidentes sin atender?",
            "Sin señales de necesidad",
            "Madurez baja / sin abogado conocido",
            "Madurez media, sin firma retenida",
            "Medio",
            "Investigar estructura"
        ),
        (
            "C6: Presencia en Medellín",
            "¿Opera en Medellín o Antioquia?",
            "Sin sede en Colombia",
            "Sede en otra ciudad colombiana",
            "Sede confirmada en Medellín / Antioquia",
            "Medio",
            "Prioridad geográfica"
        ),
        (
            "C7: Urgencia del Problema",
            "¿Tiene un problema legal activo o inminente?",
            "Sin urgencia visible",
            "Expansión/crecimiento (urgencia media)",
            "Litigio, sanción o M&A en curso",
            "Alto",
            "Señal temporal crítica"
        ),
        (
            "C8: Sin Riesgo BigLaw",
            "¿Es probable que no tenga firma grande retenida?",
            "Unicornio/multinacional con BigLaw evidente",
            "Empresa grande sin firma visible",
            "Pyme/scaleup sin firma externa evidente",
            "Medio",
            "Verificar noticias"
        ),
    ]

    for i, row_data in enumerate(criterios_data):
        row = i + 5
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        for j, v in enumerate(row_data):
            c = ws.cell(row=row, column=j+1)
            c.value = v
            c.font = cell_font(size=9, bold=(j == 0))
            c.alignment = left() if j != 5 else center()
            c.border = thin_border()
            c.fill = PatternFill("solid", fgColor=bg)
            if j == 5:
                if v == "Alto":
                    c.fill = PatternFill("solid", fgColor=C_VERDE)
                    c.font = Font(name="Calibri", bold=True, size=9, color=C_VERDE_F)
                else:
                    c.fill = PatternFill("solid", fgColor=C_AMARILLO)
                    c.font = Font(name="Calibri", bold=True, size=9, color=C_AMARILLO_F)
        ws.row_dimensions[row].height = 20

    # Tabla de semáforo
    ws.row_dimensions[14].height = 14

    ws.merge_cells("A15:G15")
    c = ws["A15"]
    c.value = "TABLA DE CLASIFICACIÓN (Semáforo)"
    c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill = header_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[15].height = 22

    semaforos = [
        ("🟢 VERDE — Prospecto A",   "33–40", "Contactar esta semana. Mensaje personalizado de alto valor.",      C_VERDE,    C_VERDE_F),
        ("🟡 AMARILLO — Prospecto B", "24–32", "Incluir en nurturing 30 días. Monitorear nuevas señales.",         C_AMARILLO, C_AMARILLO_F),
        ("🟠 NARANJA — Prospecto C",  "15–23", "Seguimiento pasivo. Conectar en LinkedIn. Revisar en 60 días.",    C_NARANJA,  C_NARANJA_F),
        ("🔴 ROJO — Descartar",       "0–14",  "No invertir tiempo. Revisar en 6 meses si aparece nueva señal.",  C_ROJO,     C_ROJO_F),
    ]

    ws.merge_cells("A16:B16")
    ws.merge_cells("C16:C16")
    ws.merge_cells("D16:G16")

    for j, hdr in enumerate(["Clasificación", "Puntaje", "Acción recomendada"]):
        cols_map = {0: 1, 1: 3, 2: 4}
        c = ws.cell(row=16, column=cols_map[j])
        c.value = hdr
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = header_fill(GOLD)
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[16].height = 18

    for i, (label, pts, accion, bg, fg) in enumerate(semaforos):
        row = 17 + i
        # Columnas A-B merged: label
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1)
        c.value = label
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Calibri", bold=True, size=10, color=fg)
        c.alignment = center()
        c.border = thin_border()
        # Columna C: puntaje
        c2 = ws.cell(row=row, column=3)
        c2.value = pts
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.font = Font(name="Calibri", bold=True, size=10, color=fg)
        c2.alignment = center()
        c2.border = thin_border()
        # Columnas D-G merged: acción
        ws.merge_cells(f"D{row}:G{row}")
        c3 = ws.cell(row=row, column=4)
        c3.value = accion
        c3.fill = PatternFill("solid", fgColor=bg)
        c3.font = Font(name="Calibri", size=9, color=fg)
        c3.alignment = left()
        c3.border = thin_border()
        ws.row_dimensions[row].height = 18

# ─── HOJA 4: MENSAJES DE CONTACTO ────────────────────────────────────────────

def crear_mensajes(wb, prospectos):
    ws = wb.create_sheet("✉️ Mensajes")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "MENSAJES DE CONTACTO — JA ABOGADOS  |  Copiar y personalizar con [Tu nombre]"
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = header_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 28

    for col, (name, width) in enumerate(zip(
        ["Empresa", "Mensaje de Conexión (LinkedIn ≤300 car.)", "Primer Mensaje (DM LinkedIn)", "Mensaje de Seguimiento"],
        [18, 40, 55, 55]
    )):
        c = ws.cell(row=2, column=col+1)
        c.value = name
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = header_fill(GOLD)
        c.alignment = center()
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col+1)].width = width
    ws.row_dimensions[2].height = 20

    FALLBACK = {
        "conexion": "Hola, sigo con interés el crecimiento de [EMPRESA] en Medellín. Me dedico al derecho corporativo y me encantaría conectar para compartir perspectivas del sector.",
        "mensaje1": "Buen día. Noté que [EMPRESA] tiene una trayectoria interesante en su sector. Desde mi práctica en derecho corporativo en Medellín acompaño a empresas en estructuración jurídica, contratos y compliance. Si en algún momento necesitan una revisión de su situación legal, con gusto les doy una primera consulta sin costo. ¿Están trabajando con apoyo jurídico externo actualmente?",
        "seguimiento": "¿Cómo les va con el crecimiento? Quería compartirles una nota breve sobre [TEMA RELEVANTE AL SECTOR] que puede ser útil para empresas en su etapa. La comparto con mucho gusto si le parece.",
    }

    for i, p in enumerate(prospectos):
        row = i + 3
        empresa = p[0]
        msgs = MENSAJES.get(empresa, {})

        bg = LIGHT_GRAY if i % 2 == 0 else WHITE

        for j, (key, fallback) in enumerate(zip(
            ["conexion", "mensaje1", "seguimiento"],
            [FALLBACK["conexion"].replace("[EMPRESA]", empresa),
             FALLBACK["mensaje1"].replace("[EMPRESA]", empresa),
             FALLBACK["seguimiento"].replace("[EMPRESA]", empresa)]
        )):
            col = j + 2
            c = ws.cell(row=row, column=col)
            c.value = msgs.get(key, fallback)
            c.font = cell_font(size=9)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = thin_border()
            c.fill = PatternFill("solid", fgColor=bg)

        # Empresa
        c = ws.cell(row=row, column=1)
        c.value = empresa
        c.font = cell_font(bold=True, size=9, color=NAVY)
        c.alignment = center()
        c.border = thin_border()
        c.fill = PatternFill("solid", fgColor=bg)

        ws.row_dimensions[row].height = 80

# ─── HOJA 5: INSTRUCCIONES ───────────────────────────────────────────────────

def crear_instrucciones(wb):
    ws = wb.create_sheet("📖 Instrucciones")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 90

    ws.merge_cells("A1:A1")
    c = ws["A1"]
    c.value = "GUÍA DE USO — CRM JA ABOGADOS"
    c.font = Font(name="Calibri", bold=True, size=16, color=WHITE)
    c.fill = header_fill(NAVY)
    c.alignment = center()
    ws.row_dimensions[1].height = 35

    instrucciones = [
        ("", ""),
        ("CÓMO SUBIR ESTE ARCHIVO A GOOGLE SHEETS", "header"),
        ("1. Abra drive.google.com en su navegador.", "paso"),
        ("2. Haga clic en '+ Nuevo' → 'Subir archivo' y seleccione crm_jaabogados.xlsx.", "paso"),
        ("3. Una vez subido, haga doble clic sobre el archivo.", "paso"),
        ("4. Google Sheets mostrará una barra amarilla — haga clic en 'Abrir con Google Sheets'.", "paso"),
        ("5. Guarde como Hoja de cálculo de Google: Archivo → Guardar como Hoja de cálculo de Google.", "paso"),
        ("6. Desde ahora trabaje siempre desde la versión en Google Sheets (no el .xlsx original).", "paso"),
        ("", ""),
        ("CÓMO USAR EL CRM (Hoja '📋 CRM Prospectos')", "header"),
        ("• Cada fila es un prospecto. Los 17 iniciales ya están precargados con sus datos y puntajes.", "item"),
        ("• Columna L (ESTADO): use el menú desplegable para actualizar el estado de cada prospecto.", "item"),
        ("• Columnas N-U (C1 a C8): si consigue más información, actualice el puntaje (0-5) de cada criterio.", "item"),
        ("• Columna V (TOTAL): suma automática de C1–C8. Si ve cero, escriba la fórmula =SUMA(N3:U3) en V3 y arrastre.", "item"),
        ("• Columna W (SEMÁFORO): ya viene calculado. Si añade prospectos nuevos, copie la fórmula de la fila anterior.", "item"),
        ("• Columna K (Fecha primer contacto): registre la fecha cuando haga el primer contacto.", "item"),
        ("• Columna X (Próxima Acción) y Y (Fecha): actualice cada vez que defina el siguiente paso.", "item"),
        ("", ""),
        ("CÓMO AGREGAR UN PROSPECTO NUEVO", "header"),
        ("1. Vaya a la última fila con datos en '📋 CRM Prospectos'.", "paso"),
        ("2. En la fila siguiente, llene los campos B a M con los datos de la empresa.", "paso"),
        ("3. Evalúe la empresa con la Matriz (hoja '🎯 Matriz de Evaluación') y llene C1-C8 (columnas N-U).", "paso"),
        ("4. Copie la fórmula de TOTAL (columna V) y SEMÁFORO (columna W) de la fila anterior.", "paso"),
        ("5. Asigne el estado 'Prospecto Nuevo' en la columna L.", "paso"),
        ("", ""),
        ("FLUJO SEMANAL RECOMENDADO (30 minutos/semana)", "header"),
        ("Lunes:   Revisar prospectos VERDE — ¿envié el mensaje? ¿hubo respuesta? Actualizar estado.", "item"),
        ("Martes:  Revisar prospectos AMARILLO — aplicar secuencia de nurturing (semana 2 o 3).", "item"),
        ("Viernes: Agregar prospectos nuevos identificados durante la semana. Evaluar con la Matriz.", "item"),
        ("Fin de mes: Revisar prospectos NARANJA — ¿alguno tuvo nueva señal? ¿subir a Amarillo?", "item"),
        ("", ""),
        ("REGLAS DE ORO (del PDF 'Tu Buscador de Clientes Potenciales')", "header"),
        ("1. Nunca inventar datos. Solo incluir información verificada en fuentes reales.", "item"),
        ("2. Verificar en internet antes de contactar. Conocer la señal específica de cada empresa.", "item"),
        ("3. Los mensajes son conversaciones, no ventas. Primero valor, luego oferta.", "item"),
        ("4. Máximo 10-15 mensajes por día en LinkedIn para no ser marcado como spam.", "item"),
        ("5. Respetar la Ley 1581/2012 (Habeas Data): no compartir esta base de datos con terceros.", "item"),
        ("", ""),
        ("SOPORTE Y ACTUALIZACIONES", "header"),
        ("Este archivo fue generado con Claude (Anthropic) para JA Abogados — Medellín, Colombia.", "item"),
        ("Versión 1.0 — Junio 2026. Para actualizar la lista de prospectos, siga el flujo de la sección 3.", "item"),
    ]

    for i, (text, tipo) in enumerate(instrucciones):
        row = i + 2
        c = ws.cell(row=row, column=1)
        c.value = text
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if tipo == "header":
            c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
            c.fill = header_fill(NAVY)
            ws.row_dimensions[row].height = 22
        elif tipo == "paso":
            c.font = Font(name="Calibri", size=10, color="000000")
            c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            ws.row_dimensions[row].height = 16
        elif tipo == "item":
            c.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            c.fill = PatternFill("solid", fgColor=WHITE)
            ws.row_dimensions[row].height = 16
        else:
            ws.row_dimensions[row].height = 8

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()

    print("Creando hoja: Dashboard...")
    crear_dashboard(wb, PROSPECTOS)

    print("Creando hoja: CRM Prospectos...")
    crear_crm(wb, PROSPECTOS)

    print("Creando hoja: Matriz de Evaluación...")
    crear_matriz(wb)

    print("Creando hoja: Mensajes...")
    crear_mensajes(wb, PROSPECTOS)

    print("Creando hoja: Instrucciones...")
    crear_instrucciones(wb)

    output = "/home/user/jacabogados/crm_jaabogados.xlsx"
    wb.save(output)
    print(f"\n✅  Archivo generado: {output}")
    print(f"    Tamaño: {__import__('os').path.getsize(output):,} bytes")
    print("\nHojas creadas:")
    for s in wb.sheetnames:
        print(f"  • {s}")
    print("\nPróximo paso: suba crm_jaabogados.xlsx a Google Drive")
    print("y ábralo con Google Sheets (ver hoja 📖 Instrucciones).")

if __name__ == "__main__":
    main()
